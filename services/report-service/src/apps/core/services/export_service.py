"""
Export Service.

Service for exporting reports to various formats.
"""
import io
import json
import csv
import logging
from typing import Dict, Any, List, Optional
from datetime import datetime
from uuid import uuid4

from django.conf import settings

from ..exceptions import ReportExportFailed
from ..constants import FORMAT_PDF, FORMAT_EXCEL, FORMAT_CSV, FORMAT_JSON, FORMAT_HTML

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting reports to various formats."""

    @classmethod
    def export(
        cls,
        data: List[Dict],
        columns: List[Dict],
        format: str,
        title: str,
        chart_type: str = "",
        visualization_config: Optional[Dict] = None,
    ) -> Dict[str, Any]:
        """
        Export data to the specified format.

        Args:
            data: Data rows to export
            columns: Column definitions
            format: Output format (pdf, excel, csv, json, html)
            title: Report title
            chart_type: Chart type if applicable
            visualization_config: Visualization settings

        Returns:
            Dict with file_url and file_size

        Raises:
            ReportExportFailed: If export fails
        """
        try:
            if format == FORMAT_CSV:
                return cls._export_csv(data, columns, title)
            elif format == FORMAT_JSON:
                return cls._export_json(data, columns, title)
            elif format == FORMAT_EXCEL:
                return cls._export_excel(data, columns, title)
            elif format == FORMAT_PDF:
                return cls._export_pdf(data, columns, title, chart_type, visualization_config)
            elif format == FORMAT_HTML:
                return cls._export_html(data, columns, title, chart_type, visualization_config)
            else:
                raise ReportExportFailed(detail=f"Unsupported format: {format}")

        except ReportExportFailed:
            raise
        except Exception as e:
            logger.error(f"Export failed: {e}", exc_info=True)
            raise ReportExportFailed(detail=str(e))

    @classmethod
    def _export_csv(cls, data: List[Dict], columns: List[Dict], title: str) -> Dict[str, Any]:
        """Export to CSV format."""
        output = io.StringIO()
        fieldnames = [col['field'] for col in columns]
        headers = {col['field']: col.get('header', col['field']) for col in columns}

        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')

        # Write headers
        writer.writerow(headers)

        # Write data
        for row in data:
            writer.writerow(row)

        content = output.getvalue()
        file_size = len(content.encode('utf-8'))

        # In production, save to storage and return URL
        file_url = cls._save_to_storage(content.encode('utf-8'), title, 'csv')

        return {
            'file_url': file_url,
            'file_size': file_size,
            'content_type': 'text/csv',
        }

    @classmethod
    def _export_json(cls, data: List[Dict], columns: List[Dict], title: str) -> Dict[str, Any]:
        """Export to JSON format."""
        # Build export structure
        export_data = {
            'title': title,
            'generated_at': datetime.utcnow().isoformat(),
            'columns': columns,
            'data': data,
            'row_count': len(data),
        }

        content = json.dumps(export_data, indent=2, default=str)
        file_size = len(content.encode('utf-8'))

        file_url = cls._save_to_storage(content.encode('utf-8'), title, 'json')

        return {
            'file_url': file_url,
            'file_size': file_size,
            'content_type': 'application/json',
        }

    @classmethod
    def _export_excel(cls, data: List[Dict], columns: List[Dict], title: str) -> Dict[str, Any]:
        """Export to Excel format."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback to CSV if openpyxl not available
            logger.warning("openpyxl not available, falling back to CSV")
            return cls._export_csv(data, columns, title)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit

        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col.get('header', col['field'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

            # Set column width
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(col['field'])
                cell.border = thin_border

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        file_size = len(content)

        file_url = cls._save_to_storage(content, title, 'xlsx')

        return {
            'file_url': file_url,
            'file_size': file_size,
            'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }

    @classmethod
    def _export_pdf(
        cls,
        data: List[Dict],
        columns: List[Dict],
        title: str,
        chart_type: str,
        visualization_config: Optional[Dict],
    ) -> Dict[str, Any]:
        """Export to PDF format."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            # Fallback to HTML if reportlab not available
            logger.warning("reportlab not available, falling back to HTML")
            return cls._export_html(data, columns, title, chart_type, visualization_config)

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20
        )

        elements = []

        # Title
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(
            f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 20))

        # Table data
        table_data = []

        # Headers
        headers = [col.get('header', col['field']) for col in columns]
        table_data.append(headers)

        # Rows
        for row in data[:500]:  # Limit rows for PDF
            row_values = [str(row.get(col['field'], '')) for col in columns]
            table_data.append(row_values)

        # Create table
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        content = output.getvalue()
        file_size = len(content)

        file_url = cls._save_to_storage(content, title, 'pdf')

        return {
            'file_url': file_url,
            'file_size': file_size,
            'content_type': 'application/pdf',
        }

    @classmethod
    def _export_html(
        cls,
        data: List[Dict],
        columns: List[Dict],
        title: str,
        chart_type: str,
        visualization_config: Optional[Dict],
    ) -> Dict[str, Any]:
        """Export to HTML format."""
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 20px; }}
        h1 {{ color: #333; }}
        .meta {{ color: #666; margin-bottom: 20px; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
        th {{ background-color: #4472C4; color: white; padding: 12px 8px; text-align: left; }}
        td {{ padding: 10px 8px; border-bottom: 1px solid #ddd; }}
        tr:nth-child(even) {{ background-color: #f8f9fa; }}
        tr:hover {{ background-color: #e8f4ff; }}
    </style>
</head>
<body>
    <h1>{title}</h1>
    <p class="meta">Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Rows: {len(data)}</p>
    <table>
        <thead>
            <tr>
                {''.join(f'<th>{col.get("header", col["field"])}</th>' for col in columns)}
            </tr>
        </thead>
        <tbody>
"""
        for row in data:
            html += "<tr>"
            for col in columns:
                value = row.get(col['field'], '')
                html += f"<td>{value}</td>"
            html += "</tr>\n"

        html += """
        </tbody>
    </table>
</body>
</html>"""

        content = html.encode('utf-8')
        file_size = len(content)

        file_url = cls._save_to_storage(content, title, 'html')

        return {
            'file_url': file_url,
            'file_size': file_size,
            'content_type': 'text/html',
        }

    @classmethod
    def _save_to_storage(cls, content: bytes, title: str, extension: str) -> str:
        """
        Save content to storage and return URL.

        In production, this would save to S3/GCS/Azure Blob.
        For now, returns a mock URL.
        """
        # Generate unique filename
        safe_title = "".join(c if c.isalnum() else "_" for c in title)[:50]
        filename = f"{safe_title}_{uuid4().hex[:8]}.{extension}"

        # In production: save to storage service
        # storage_client.upload(content, filename)

        base_url = getattr(settings, 'REPORTS_BASE_URL', '/api/v1/reports/files')
        return f"{base_url}/{filename}"
